"""Create a clean template from the contoh Excel by clearing data cells, keeping formulas."""
import openpyxl, shutil
from pathlib import Path

SRC = Path(__file__).parent.parent / "Contoh_Working Report - Rakarizal Muhammad Zidan - Jun 2026.xlsx"
DST = Path(__file__).parent / "template.xlsx"

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)
ws = wb.active

# Clear header info (keep labels, clear values)
ws["E3"].value = None  # Employee Name
ws["E4"].value = None  # ID No
ws["E5"].value = None  # Position
ws["K3"].value = None  # Customer Name (label in J3 stays)
ws["K4"].value = None  # Project Name
ws["K5"].value = None  # WO Number
ws["E7"].value = None  # Month
ws["F7"].value = None  # Year
ws["F44"].value = None  # Issued by name
ws["F45"].value = None  # Issued date

# Clear data rows 10-40: only clear value cells (E, F, G, J, K), leave formula cells alone
for row in range(10, 41):
    for col_letter in ("E", "F", "G", "J", "K"):
        ws[f"{col_letter}{row}"].value = None

wb.save(DST)
print(f"Clean template saved: {DST}")
