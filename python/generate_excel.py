"""Generate a Working Report Excel from SQLite data and the company template.

Usage:
    python generate_excel.py                    # previous month
    python generate_excel.py --month 6 --year 2026
    python generate_excel.py --db path/to/worklog.db
"""
import argparse
import datetime
import os
import shutil
import sqlite3
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).parent
DEFAULT_DB = SCRIPT_DIR.parent / "database" / "worklog.db"
TEMPLATE = SCRIPT_DIR / "template.xlsx"
OUTPUT_DIR = SCRIPT_DIR.parent / "output"


def parse_time(text):
    """Convert 'HH:MM' or 'H:MM' string to datetime.time, or None."""
    if not text:
        return None
    try:
        parts = text.strip().split(":")
        return datetime.time(int(parts[0]), int(parts[1]))
    except (ValueError, IndexError):
        return None


def parse_break(text):
    """Convert break_time text to datetime.time.

    Handles: '1 Hour', '01:00', '1:00', '0:30', etc.
    """
    if not text:
        return datetime.time(1, 0)
    t = text.strip().lower()
    if "hour" in t:
        # ponytail: naive — assumes '1 Hour' or '2 Hour'. upgrade: regex.
        hrs = int("".join(c for c in t if c.isdigit()) or "1")
        return datetime.time(hrs, 0)
    return parse_time(text) or datetime.time(1, 0)


def load_env():
    """Load .env file if it exists. stdlib only, no dotenv dependency."""
    env_path = SCRIPT_DIR.parent / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, val = line.partition("=")
        os.environ.setdefault(key.strip(), val.strip())


def generate(month: int, year: int, db_path: Path):
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE}. Run create_template.py first.")
    if not db_path.exists():
        raise FileNotFoundError(f"Database not found: {db_path}. Run database/init_db.py first.")

    load_env()

    # Query work logs for the month
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM work_logs WHERE strftime('%Y', work_date) = ? AND strftime('%m', work_date) = ? ORDER BY work_date",
        (str(year), f"{month:02d}"),
    ).fetchall()
    conn.close()

    # Build lookup: day_of_month -> row data
    logs_by_day = {}
    for r in rows:
        day = int(r["work_date"].split("-")[2])
        # ponytail: last-write-wins if multiple entries per day. upgrade: concatenate activities.
        logs_by_day[day] = r

    # Copy template
    OUTPUT_DIR.mkdir(exist_ok=True)
    month_name = datetime.date(year, month, 1).strftime("%b")
    employee_name = os.environ.get("EMPLOYEE_NAME", "Employee")
    out_filename = f"Working Report - {employee_name} - {month_name} {year}.xlsx"
    out_path = OUTPUT_DIR / out_filename
    shutil.copy2(TEMPLATE, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # Fill header
    ws["E7"] = month
    ws["F7"] = year
    ws["E3"] = employee_name
    ws["E4"] = os.environ.get("EMPLOYEE_ID", "")
    ws["E5"] = os.environ.get("EMPLOYEE_POSITION", "")
    ws["K3"] = os.environ.get("CUSTOMER_NAME", "")
    ws["K4"] = os.environ.get("PROJECT_NAME", "")
    ws["K5"] = os.environ.get("WO_NUMBER", "")
    ws["F44"] = employee_name
    ws["F45"] = datetime.datetime(year, month, 1)

    # Fill data rows: row 10 = day 1, row 11 = day 2, ..., row 40 = day 31
    filled = 0
    for day in range(1, 32):
        excel_row = 9 + day  # day 1 -> row 10
        if day not in logs_by_day:
            continue
        log = logs_by_day[day]

        activity_text = log["ai_activity"] or log["activity"]
        check_in = parse_time(log["check_in"])
        check_out = parse_time(log["check_out"])
        break_t = parse_break(log["break_time"])

        ws.cell(row=excel_row, column=5).value = check_in    # E: IN
        ws.cell(row=excel_row, column=6).value = check_out   # F: OUT
        ws.cell(row=excel_row, column=7).value = break_t     # G: BREAK
        ws.cell(row=excel_row, column=10).value = log["place"]  # J: Place
        ws.cell(row=excel_row, column=11).value = activity_text  # K: Activity
        filled += 1

    wb.save(out_path)
    print(f"Generated: {out_path} ({filled} days filled from {len(rows)} records)")
    return out_path


def main():
    today = datetime.date.today()
    # Default: previous month
    first_of_this_month = today.replace(day=1)
    prev_month = first_of_this_month - datetime.timedelta(days=1)

    parser = argparse.ArgumentParser(description="Generate Working Report Excel")
    parser.add_argument("--month", type=int, default=prev_month.month)
    parser.add_argument("--year", type=int, default=prev_month.year)
    parser.add_argument("--db", type=str, default=str(DEFAULT_DB))
    args = parser.parse_args()

    out = generate(args.month, args.year, Path(args.db))

    # ponytail: self-check — verify output is valid xlsx with expected header
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws["E7"].value == args.month, f"Month mismatch: {ws['E7'].value}"
    assert ws["F7"].value == args.year, f"Year mismatch: {ws['F7'].value}"
    print("Self-check passed.")


if __name__ == "__main__":
    main()
